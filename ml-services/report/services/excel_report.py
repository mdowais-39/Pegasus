"""Render the report data model to an .xlsx workbook (Core Requirement 6).

Minimal, statistics-first and investigator-friendly. No charts — just the
numbers (on the Summary sheet) plus the two explained sheets (round trips and
money flow). Every account identifier is already resolved to an account number /
file name upstream in report_builder, so no opaque statement UUID appears here.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_SECTION_FONT = Font(bold=True, size=12, color="1F4E78")
_WRAP = Alignment(wrap_text=True, vertical="top")


def _money(v):
    try:
        return f"Rs {float(v):,.0f}"
    except (TypeError, ValueError):
        return "-" if v in (None, "") else str(v)


def _sheet(wb, title):
    return wb.create_sheet(title=title[:31])


def _header_row(ws, headers, row=1):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL


def _autosize(ws, max_width=70):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_width, width + 2)


KEY_METRICS = [
    ("statements", "Statements"),
    ("transactions", "Transactions"),
    ("total_credit", "Total Credit"),
    ("total_debit", "Total Debit"),
    ("duplicates", "Duplicate Transactions"),
    ("failed_or_reversed", "Failed / Reversed"),
    ("round_trips_detected", "Round Trips Detected"),
    ("high_risk_accounts", "High-Risk Accounts"),
]
_MONEY_METRICS = {"total_credit", "total_debit"}

CATEGORY_LABELS = [
    ("total_transactions", "Total Transactions"), ("credits", "Credits"),
    ("debits", "Debits"), ("atm_withdrawals", "ATM Withdrawals"),
    ("cash_deposits", "Cash Deposits"), ("failed_transactions", "Failed / Reversed"),
    ("digital_upi", "Digital (UPI / apps)"), ("cheque", "Cheque"),
    ("card_pos", "Card / POS"),
]


def _summary_sheet(wb, report):
    """Every statistic in one place — key metrics, risk mix, categories,
    channels — as plain number tables (no charts)."""
    ws = _sheet(wb, "Summary")
    ws["A1"] = report.get("title", "Investigation Report")
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = report.get("scope") or f"Case: {report.get('case_id', 'all')}"
    if report.get("generated_at"):
        ws["A3"] = f"Generated: {report.get('generated_at')}"

    r = 5

    def section(title):
        nonlocal r
        ws.cell(row=r, column=1, value=title).font = _SECTION_FONT
        r += 1

    def head(cols):
        nonlocal r
        _header_row(ws, cols, row=r)
        r += 1

    # Key metrics
    section("Key Metrics")
    head(["Metric", "Value"])
    es = report.get("executive_summary", {}) or {}
    for key, lbl in KEY_METRICS:
        val = _money(es.get(key)) if key in _MONEY_METRICS else es.get(key, 0)
        ws.cell(row=r, column=1, value=lbl)
        ws.cell(row=r, column=2, value=val)
        r += 1
    r += 1

    # Risk distribution
    dist = report.get("risk_distribution") or {}
    if any(dist.values()):
        section("Accounts by Risk Level")
        head(["Risk Level", "Accounts"])
        for lvl in ("CRITICAL", "HIGH", "MEDIUM", "LOW"):
            ws.cell(row=r, column=1, value=lvl.title())
            ws.cell(row=r, column=2, value=dist.get(lvl, 0))
            r += 1
        r += 1

    # Transaction categories
    cats = report.get("category_counts") or {}
    if cats:
        section("Transaction Categories")
        head(["Category", "Count"])
        for key, lbl in CATEGORY_LABELS:
            ws.cell(row=r, column=1, value=lbl)
            ws.cell(row=r, column=2, value=cats.get(key, 0))
            r += 1
        r += 1

    # Payment channels
    channels = report.get("channel_breakdown") or []
    if channels:
        section("Payment Channels")
        head(["Channel / Class", "Transactions", "Total Value", "Share"])
        for c in channels:
            ws.cell(row=r, column=1, value=c.get("channel"))
            ws.cell(row=r, column=2, value=c.get("count"))
            ws.cell(row=r, column=3, value=_money(c.get("value")))
            ws.cell(row=r, column=4, value=f"{round((c.get('share') or 0) * 100)}%")
            r += 1
        r += 1

    # Flagged findings — kept on the summary sheet (not a separate sheet)
    section("Flagged Findings")
    ws.cell(row=r, column=1, value=report.get("flags_summary", ""))
    r += 1
    head(["Account", "Severity", "Flags", "Evidence"])
    for f in report.get("flagged_findings", []):
        ws.cell(row=r, column=1, value=f.get("account"))
        ws.cell(row=r, column=2, value=f.get("severity"))
        ws.cell(row=r, column=3, value=", ".join(f.get("tags", []) or []))
        c = ws.cell(row=r, column=4, value="; ".join(f.get("reasons", []) or []))
        c.alignment = _WRAP
        r += 1

    _autosize(ws)
    ws.column_dimensions["D"].width = 55


def _round_trips_sheet(wb, report):
    ws = _sheet(wb, "Round Trips")
    ws["A1"] = "Round Trips (Circular Flows)"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = report.get("round_trips_definition", "")
    ws["A2"].alignment = _WRAP
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 60

    _header_row(ws, ["Chain", "Path", "Min Amount", "Total Amount", "What it means"], row=4)
    r = 5
    for i, c in enumerate(report.get("round_trips", []), start=1):
        ws.cell(row=r, column=1, value=f"#{c.get('id', i)}")
        p = ws.cell(row=r, column=2, value=" -> ".join(str(n) for n in c.get("nodes", [])))
        p.alignment = _WRAP
        ws.cell(row=r, column=3, value=_money(c.get("min_amount")))
        ws.cell(row=r, column=4, value=_money(c.get("total_amount")))
        d = ws.cell(row=r, column=5, value=c.get("description", ""))
        d.alignment = _WRAP
        r += 1
    _autosize(ws)
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["E"].width = 55


def _money_flow_sheet(wb, report):
    ws = _sheet(wb, "Money Flow")
    defs = report.get("money_flow_definitions") or {}
    mf = report.get("money_flow", {})
    ws["A1"] = "Money Flow"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = defs.get("overview", "")
    ws["A2"].alignment = _WRAP
    ws.merge_cells("A2:C2")
    ws.row_dimensions[2].height = 45
    ws["A4"] = "Primary destination (where funds accumulate):"
    ws["A4"].font = Font(bold=True)
    ws["B4"] = mf.get("destination_account") or "None identified"

    r = 6

    def block(title, definition, headers, rows):
        nonlocal r
        if not rows:
            return
        ws.cell(row=r, column=1, value=title).font = _SECTION_FONT
        r += 1
        if definition:
            c = ws.cell(row=r, column=1, value=definition)
            c.alignment = _WRAP
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
            ws.row_dimensions[r].height = 45
            r += 1
        _header_row(ws, headers, row=r)
        r += 1
        for row in rows:
            for ci, val in enumerate(row, start=1):
                ws.cell(row=r, column=ci, value=val)
            r += 1
        r += 1

    block("Accumulation Accounts", defs.get("accumulation"),
          ["Account", "Total Received", "Senders"],
          [[a.get("node"), _money(a.get("total_received")), a.get("sender_count")]
           for a in mf.get("accumulation_accounts", [])[:10]])
    block("Source Accounts", defs.get("source"),
          ["Account", "Total Sent", "Receivers"],
          [[a.get("node"), _money(a.get("total_sent")), a.get("receiver_count")]
           for a in mf.get("source_accounts", [])[:10]])
    block("Layering / Pass-Through Accounts", defs.get("layering"),
          ["Account", "Total In", "Total Out", "Pass-Through"],
          [[a.get("node"), _money(a.get("total_in")), _money(a.get("total_out")),
            f"{round((a.get('passthrough_ratio') or 0) * 100)}%"]
           for a in mf.get("layering", [])[:10]])
    _autosize(ws)


def _fund_velocity_sheet(wb, report):
    timeline = report.get("activity_timeline") or []
    if not timeline:
        return
    ws = _sheet(wb, "Fund Velocity")
    ws["A1"] = "Fund Velocity - Daily Activity"
    ws["A1"].font = _TITLE_FONT
    _header_row(ws, ["Date", "Transactions", "Credit", "Debit"], row=3)
    r = 4
    for t in timeline:
        ws.cell(row=r, column=1, value=t.get("date"))
        ws.cell(row=r, column=2, value=t.get("count"))
        ws.cell(row=r, column=3, value=_money(t.get("credit")))
        ws.cell(row=r, column=4, value=_money(t.get("debit")))
        r += 1
    _autosize(ws)


def _cash_locations_sheet(wb, report):
    cash = report.get("cash_locations") or []
    if not cash:
        return
    ws = _sheet(wb, "Cash Locations")
    ws["A1"] = "Cash Withdrawal / Deposit Locations"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = ("Physical ATM / branch locations where cash was withdrawn or "
                "deposited, parsed from the transaction narrations.")
    ws["A2"].alignment = _WRAP
    ws.merge_cells("A2:F2")
    _header_row(ws, ["City", "State", "Type", "Amount", "Date", "Time"], row=4)
    r = 5
    for c in cash:
        ws.cell(row=r, column=1, value=c.get("city"))
        ws.cell(row=r, column=2, value=c.get("state"))
        ws.cell(row=r, column=3,
                value="Withdrawal" if c.get("direction") == "DEBIT" else "Deposit")
        ws.cell(row=r, column=4, value=_money(c.get("amount")))
        ws.cell(row=r, column=5, value=c.get("date"))
        ws.cell(row=r, column=6, value=c.get("time"))
        r += 1
    _autosize(ws)


def build_excel(report: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    _summary_sheet(wb, report)       # incl. flagged findings
    _round_trips_sheet(wb, report)
    _money_flow_sheet(wb, report)
    _fund_velocity_sheet(wb, report)
    _cash_locations_sheet(wb, report)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
